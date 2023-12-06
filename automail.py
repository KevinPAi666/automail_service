### get api service and .json -> https://console.developers.google.com/apis/
html_biglogo = "./automail/pic_biglogo.png"
html_signlogo = "./automail/pic_signlogo.png"
html_card_check = "./automail/automail_check.html"
html_card_complete = "./automail/automail_complete.html"

# important !!!
# when setting finished, shouldn't modify again.
myAccount = "XXX" #enter gmail account who sent in here
myPassword = "XXX" #enter gmail application password who sent in here
json_file = "./XXX" #XXX = your json file name

# customize course setting
### test form  -> enter your sheet link for convenience
google_form_backend = "XXX" #enter your google sheet backend linkin hrer

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email import policy
from openpyxl import load_workbook
import smtplib
import requests, requests_oauthlib
import gspread
import time, datetime, os


def course_deadline_process(course_deadline):
    _weekd = ['一', '二', '三', '四', '五', '六', '日']
    _paytime = datetime.datetime.now() + datetime.timedelta(days=3)
    deadline = datetime.datetime.strptime(course_deadline, "%Y-%m-%d")
    if _paytime <= deadline :
        paytime = _paytime
    else:
        paytime = deadline
    paytime_weekd = _weekd[paytime.weekday()]
    paytime = str(paytime).split(' ')[0]
    return paytime, paytime_weekd

def get_course_info(course_name):
    form = "./automail/course_overall.xlsx"
    excel = load_workbook(form).active
    total_course = excel.max_row + 1
    course_deadline, course_link = 0, 0
    for i in range(2, total_course):
        if excel['A'+str(i)].value == course_name:
            course_deadline = str(excel['B'+str(i)].value).split(' ')[0]
            course_link = excel['C'+str(i)].value
            course_price = excel['D'+str(i)].value
    return course_deadline, course_link, course_price

# start in here
def get_forms():
    print('>>>>>>>>>>>>>>>>>>>>>>>>')
    #print(os.getcwd())
    creds = gspread.service_account(filename=json_file)
    client = creds.open_by_url(google_form_backend)
    sheet = client.get_worksheet(0)
    data = sheet.get_all_values()

    #print(data)
    if "確認信寄出" not in data[0] : 
        sheet.insert_cols(values=[["確認信寄出"]], col=len(data[0])+1)
        sheet = client.get_worksheet(0)
        data = sheet.get_all_values()
        
    if "確認匯款完成" not in data[0] : 
        sheet.insert_cols(values=[["確認匯款完成"]], col=len(data[0])+1)
        sheet = client.get_worksheet(0)
        data = sheet.get_all_values()
    sent_check = len(data[0])-2
    remit_complete = len(data[0])-1

    for i in range(len(data)-1,0,-1):
        # 問題順序
        ALL_DONE = False
        global sent_fail, something_wrong
        sent_fail, something_wrong = False, False
        mailbox = data[i][1]
        name = data[i][2]
        phone = data[i][3]
        # if sheet have only one course, just fill course name, or create a new question for courses optional and get via variable, it'll contact with local excel.
        course_name = "智能辦公 - 生成式 AI 實務應用與案例分析"
        #course_name = data[i][2]

        # 確認信
        if len(data[i][sent_check]) <= 1 and len(data[i][0]) > 2: #確認提交時間不為空及確認信未寄出
            course_deadline, course_link, course_price = get_course_info(course_name)
            if course_deadline == 0 or course_link == 0:
                something_wrong = True
            if something_wrong == False:
                paytime, paytime_weekd = course_deadline_process(course_deadline)
                print(">>", name, mailbox, phone, course_name, end='\n')
                auto_send(ALL_DONE, name, mailbox, course_name, paytime, paytime_weekd, course_link, course_price)
                if not sent_fail:
                    sheet.update_acell("{}{}".format(chr(sent_check + 65), i + 1), '已完成')
            else:
                sheet.update_acell("{}{}".format(chr(sent_check + 65), i + 1), '資料異常')
                

        #付款完成信
        confirm_users = ['Eric確認', 'Ken確認']
        if data[i][remit_complete] in confirm_users and "繳費+寄送完成" not in data[i][remit_complete] and len(data[i][0]) > 2: #確認提交時間不為空及付款完成格是否被填入"XX確認"
            confirm_user = data[i][remit_complete]
            course_deadline, course_link, course_price = get_course_info(course_name)
            if course_deadline == 0 or course_link == 0:
                something_wrong = True
            if something_wrong == False:
                print(">>>>", name, mailbox, phone, course_name, confirm_user, end='\n')
                ALL_DONE = True
                auto_send(ALL_DONE, name, mailbox, course_name, course_link)
                if not sent_fail:
                    sheet.update_acell("{}{}".format(chr(remit_complete + 65), i + 1), confirm_user+'繳費+寄送完成')
        #print(name, mailbox, course_name, end='\n')


def auto_send(ALL_DONE, *args):
    content = MIMEMultipart() #policy=policy.default  #建立MIMEMultipart物件
    content["subject"] = "課程報名通知 - 「" + args[2]+'」' #郵件標題
    content["from"] = 'AI.FREE Team' #寄件者 name or mail
    content["to"] = args[1] #收件者
    content["CC"] = 'ai.free.team@gmail.com' #副本
    if ALL_DONE :
        with open(html_card_complete, 'r', encoding="UTF-8") as rr:
                 _html_content = rr.read()
        content.attach(MIMEText(_html_content.format(
                                                                                    name = args[0],
                                                                                    course_name = args[2],
                                                                                    course_link = args[3],
                                                                                    ), 'html', 'UTF-8'))  #郵件內容 html = local var
    else:
        with open(html_card_check, 'r', encoding="UTF-8") as rr:
                 _html_content = rr.read()
        content.attach(MIMEText(_html_content.format(
                                                                                    name = args[0],
                                                                                    course_name = args[2],
                                                                                    course_link = args[5],
                                                                                    course_price = args[6],
                                                                                    paytime = args[3],
                                                                                    paytime_weekd = args[4],
                                                                                    ), 'html', 'UTF-8'))  #郵件內容 html = local var
    with open(html_signlogo, 'rb') as rrr:
        _html_logo = rrr.read()
        _html_logo = MIMEImage(_html_logo)
        _html_logo.add_header('Content-ID', '<image_signlogo>')
        content.attach(_html_logo)

    with smtplib.SMTP(host="smtp.gmail.com", port="587") as smtp:  # 設定SMTP伺服器
        try:
            smtp.ehlo()  # 驗證SMTP伺服器
            smtp.starttls()  # 建立加密傳輸
            smtp.login(myAccount, myPassword)  # 登入寄件者gmail
            smtp.send_message(content)  # 寄送郵件
            print("Complete!")
        except Exception as e:
            global sent_fail
            sent_fail = True
            print("Error message: ", e)


get_forms()
print('<<<<<<<<<<<<<<<<<<<<<<<<')
